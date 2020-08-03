import os,asyncio,threading,logging,datetime,sys,datetime,time
import openpyxl
from telethon.sync import TelegramClient,events
from telethon import functions,types,utils,errors
from threading import Thread
from config import api_id, api_hash
#------------------API Credentials-----------------------
''' Need To Change These Credentials Before Deploying Project'''

#--------------------------------------------------------


#---------------Global Variables-------------------------
phone_objects = dict()
phone_status = dict()
excel_file = (os.getcwd()+"/mobile_no_samples.xlsx")
#--------------------------------------------------------



class phone:
    def __init__(self, number, wait_state,client):
        self.w = wait_state #is_waiting_or_not
        self.number = number
        self.client = client #client_connection_object


def reset_waiting_state(phone_client,waiting_time):
    global phone_objects
    global phone_status
    #print('\nThread Created for waiting State')
    time.sleep(waiting_time)
    phone_client.w = 0
    phone_status.update({phone_client.number:0})
    phone_objects.update({phone_client.number:phone(phone_client.number,0,phone_client.client)})
    #print('\nChanged object state/phone object')


def check_client_waiting():
    global phone_objects
    global phone_status
    #print('\nChecking Client Waiting Status')
    while 1:
        Free = sum(x == 0 for x in phone_status.values())
        #print('\nFree Clients are: ',Free)
        if int(Free) >= 1:
            #print('Getting Free Clients')
            for key, value in phone_status.items():
                if value == 0:
                    #print(key)
                    #print(phone_objects)
                    #print(phone_status)
                    return phone_objects[key]
                    break
        else:
            #print('all_clients are waiting')
            time.sleep(2)




phone_list = [#add Contacts here with comma seperated values and valid country extension]

loop = asyncio.get_event_loop()

async def main():
    global phone_objects
    global phone_status
    global excel_file
    start_time = datetime.datetime.now()

    #-----------Create Phone objects ------------------
    try:
        for i in phone_list:
            print('\nConnecting to - ',i)
            client = await TelegramClient(i,api_id,api_hash).start(i)
            phone_objects.update({i:phone(i,0,client)})
            phone_status.update({i:0})
        print(phone_status)
        print(phone_objects)
    except Exception as e:
        print(e)
        print(sys.exc_info()[0])
    #---------------------------------------------------
    try:
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        row_value = 1
        col_value = 1
        contacts_fetched =0
        exist = 1
        while sheet.cell(row=row_value,column=1).value and exist==True:
            ph_number = str("+"+str(sheet.cell(row=row_value,column=1).value))
            print('Checking: ',ph_number)
            status = False
            phone_client = check_client_waiting()
            await phone_client.client.connect()
            phone_client.client.flood_sleep_threshold = 0
            try:
                #print('\nFetching Entity Using: ',phone_client.number)
                result = await phone_client.client.get_entity(ph_number)
                #print(result.stringify())
                print('Exist')
                status = True
            except ValueError:
                print('Doesnot Exist')
            except errors.FloodWaitError as e:
                phone_status.update({phone_client.number:1})
                phone_objects.update({phone_client.number:phone(phone_client.number,1,phone_client)})
                #print('\nAt Main: ',phone_status)
                thread = threading.Thread(target=reset_waiting_state,args=(phone_client,e.seconds),daemon=True)
                thread.start()
            sheet.cell(row=row_value,column=2).value = str(status)
            row_value+=1
            contacts_fetched +=1
        wb.save(excel_file)
        #wb.close()
    except Exception as e:
        print(e)
        print(sys.exc_info()[0])
    finally:
        wb.save(excel_file)
        wb.close()
        end_time = datetime.datetime.now()
        print("\n---------:: Calculating Data ::------------")
        print("\nTotal_time_taken: ",end_time-start_time)
        print("\nTotal Contacts Fetched: ", contacts_fetched)



loop.run_until_complete(main())
